from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook, Workbook
import time
import os
import pandas as pd
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
# -------------------------- 请根据你的实际情况修改以下配置 --------------------------
# 1. 网页相关配置
query_url = "http://27.150.22.198:9001/login.aspx"  # 成绩查询页面的URL
# 输入框元素定位（按F12检查网页，替换为实际的name/id/xpath）
exam_num_locator = (By.NAME, "ks_ksno")    # 考生号输入框定位（示例：name属性为examNo）
ticket_num_locator = (By.NAME, "zkzh")  # 准考证号输入框定位
name_locator = (By.NAME, "ks_xm")    # 姓名输入框定位
query_button_locator = (By.XPATH, 'Button1')  # 查询按钮定位

# 2. Excel文件配置
input_excel_path = "考生信息.xlsx"    # 输入文件路径（存放待查询考生信息）
output_excel_path = "成绩结果.xlsx"   # 输出文件路径（存放查询到的成绩）
input_sheet_name = "Sheet1"           # 输入文件的工作表名
output_sheet_name = "成绩汇总"         # 输出文件的工作表名

# 3. 成绩提取配置（重点：区分有/无折算分的科目，根据实际页面HTML修改XPath）
# 科目结构：key=科目名，value=字段字典（无折算分的科目不含"折算分"键）
subjects = {
    # 无折算分的科目：语文、数学、英语、体育
    "语文": {
        "原始分": '//tr[contains(td, "语文")]/td[2]',  # 替换为实际XPath
        "等级": '//tr[contains(td, "语文")]/td[3]'     # 替换为实际XPath
    },
    "数学": {
        "原始分": '//tr[contains(td, "数学")]/td[2]',
        "等级": '//tr[contains(td, "数学")]/td[3]'
    },
    "英语": {
        "原始分": '//tr[contains(td, "英语")]/td[2]',
        "等级": '//tr[contains(td, "英语")]/td[3]'
    },
    "体育": {
        "原始分": '//tr[contains(td, "体育")]/td[2]',  # 替换为实际XPath
        "等级": '//tr[contains(td, "体育")]/td[3]'     # 替换为实际XPath
    },
    # 有折算分的科目：物理、化学、道德与法治、历史、地理、生物
    "物理": {
        "原始分": '//tr[contains(td, "物理")]/td[2]',
        "折算分": '//tr[contains(td, "物理")]/td[3]',  # 替换为实际XPath
        "等级": '//tr[contains(td, "物理")]/td[4]'
    },
    "化学": {
        "原始分": '//tr[contains(td, "化学")]/td[2]',
        "折算分": '//tr[contains(td, "化学")]/td[3]',
        "等级": '//tr[contains(td, "化学")]/td[4]'
    },
    "道德与法治": {
        "原始分": '//tr[contains(td, "道德与法治")]/td[2]',
        "折算分": '//tr[contains(td, "道德与法治")]/td[3]',
        "等级": '//tr[contains(td, "道德与法治")]/td[4]'
    },
    "历史": {
        "原始分": '//tr[contains(td, "历史")]/td[2]',
        "折算分": '//tr[contains(td, "历史")]/td[3]',
        "等级": '//tr[contains(td, "历史")]/td[4]'
    },
    "地理": {
        "原始分": '//tr[contains(td, "地理")]/td[2]',
        "折算分": '//tr[contains(td, "地理")]/td[3]',
        "等级": '//tr[contains(td, "地理")]/td[4]'
    },
    "生物": {
        "原始分": '//tr[contains(td, "生物")]/td[2]',
        "折算分": '//tr[contains(td, "生物")]/td[3]',
        "等级": '//tr[contains(td, "生物")]/td[4]'
    }
}
total_score_xpath = '//div[contains(text(), "总分")]/following-sibling::div'  # 总分的XPath（替换为实际路径）
# --------------------------------------------------------------------------------


def read_student_info(input_path, sheet_name):
    """读取输入Excel中的考生信息"""
    wb = load_workbook(input_path)
    ws = wb[sheet_name]
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:  # 跳过空行
            students.append({
                "考生号": row[0],
                "准考证号": row[1],
                "姓名": row[2]
            })
    wb.close()
    print(f"成功读取 {len(students)} 条考生信息")
    return students


def init_output_excel(output_path, sheet_name, subjects):
    """初始化输出Excel（动态生成表头：区分有/无折算分的科目）"""
    wb = Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    # 表头：考生信息 + 各科成绩 + 总分
    headers = ["考生号", "准考证号", "姓名"]
    for sub, fields in subjects.items():
        # 按科目字段动态添加表头（无折算分的科目不添加"折算分"）
        headers.append(f"{sub}_原始分")
        if "折算分" in fields:
            headers.append(f"{sub}_折算分")
        headers.append(f"{sub}_等级")
    headers.append("总分")
    ws.append(headers)
    wb.save(output_path)
    wb.close()
    print(f"已创建输出文件：{output_path}")


def extract_score(driver, subjects, total_xpath):
    """从查询结果页提取成绩（只提取科目包含的字段）"""
    score_data = {}
    # 提取各科成绩
    for sub, fields in subjects.items():
        # 提取原始分
        try:
            score_data[f"{sub}_原始分"] = driver.find_element(By.XPATH, fields["原始分"]).text.strip()
        except:
            score_data[f"{sub}_原始分"] = "未找到"
        # 若有折算分，提取折算分
        if "折算分" in fields:
            try:
                score_data[f"{sub}_折算分"] = driver.find_element(By.XPATH, fields["折算分"]).text.strip()
            except:
                score_data[f"{sub}_折算分"] = "未找到"
        # 提取等级
        try:
            score_data[f"{sub}_等级"] = driver.find_element(By.XPATH, fields["等级"]).text.strip()
        except:
            score_data[f"{sub}_等级"] = "未找到"
    # 提取总分
    try:
        score_data["总分"] = driver.find_element(By.XPATH, total_xpath).text.strip()
    except:
        score_data["总分"] = "未找到"
    return score_data


def save_to_excel(output_path, sheet_name, student, score_data, subjects):
    """将考生信息和成绩写入输出Excel（匹配动态表头）"""
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    # 组装一行数据：考生信息 + 各科成绩 + 总分
    row_data = [
        student["考生号"],
        student["准考证号"],
        student["姓名"]
    ]
    for sub, fields in subjects.items():
        # 添加原始分
        row_data.append(score_data[f"{sub}_原始分"])
        # 若有折算分，添加折算分
        if "折算分" in fields:
            row_data.append(score_data[f"{sub}_折算分"])
        # 添加等级
        row_data.append(score_data[f"{sub}_等级"])
    row_data.append(score_data["总分"])
    ws.append(row_data)
    wb.save(output_path)
    wb.close()


def main():
    # 1. 读取考生信息
    students = read_student_info(input_excel_path, input_sheet_name)
    if not students:
        print("没有找到考生信息，程序退出")
        return

    # 2. 初始化输出Excel
    if not os.path.exists(output_excel_path):
        init_output_excel(output_excel_path, output_sheet_name, subjects)

    # 3. 初始化浏览器
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))  # 使用webdriver-manager自动管理驱动
    except ConnectionError:
        print("错误：无法连接到网络下载ChromeDriver\n请检查网络连接或手动下载驱动：")
        print("1. 访问 https://sites.google.com/chromium.org/driver/ 下载对应版本ChromeDriver")
        print("2. 将以下代码替换为实际路径：")
        print("   driver = webdriver.Chrome(executable_path='C:/path/to/chromedriver.exe')")
        return
    driver.maximize_window()
    wait = WebDriverWait(driver, 10)  # 显式等待最长10秒

    try:
        # 4. 循环处理每个考生
        for i, student in enumerate(students, 1):
            print(f"\n处理第 {i}/{len(students)} 位考生：{student['姓名']}")
            try:
                driver.get(query_url)
                time.sleep(1)

                # 填充考生信息
                wait.until(EC.presence_of_element_located(exam_num_locator)).send_keys(student["考生号"])
                wait.until(EC.presence_of_element_located(ticket_num_locator)).send_keys(student["准考证号"])
                wait.until(EC.presence_of_element_located(name_locator)).send_keys(student["姓名"])
                print("已填充考生信息")

                # 点击查询
                wait.until(EC.element_to_be_clickable(query_button_locator)).click()
                print("已点击查询按钮，等待结果...")
                time.sleep(2)  # 等待结果加载

                # 提取成绩
                score_data = extract_score(driver, subjects, total_score_xpath)
                print(f"提取成绩成功：{student['姓名']} 总分：{score_data['总分']}")

                # 保存到Excel
                save_to_excel(output_excel_path, output_sheet_name, student, score_data, subjects)
                print("成绩已保存到输出文件")

            except TimeoutException:
                print(f"超时错误：页面加载过慢或元素未找到")
                # 记录错误信息
                error_data = {}
                for sub, fields in subjects.items():
                    error_data[f"{sub}_原始分"] = "超时"
                    if "折算分" in fields:
                        error_data[f"{sub}_折算分"] = "超时"
                    error_data[f"{sub}_等级"] = "超时"
                error_data["总分"] = "超时"
                save_to_excel(output_excel_path, output_sheet_name, student, error_data, subjects)
            except Exception as e:
                print(f"处理失败：{str(e)}")
                # 记录错误信息
                error_data = {}
                for sub, fields in subjects.items():
                    error_data[f"{sub}_原始分"] = "失败"
                    if "折算分" in fields:
                        error_data[f"{sub}_折算分"] = "失败"
                    error_data[f"{sub}_等级"] = "失败"
                error_data["总分"] = "失败"
                save_to_excel(output_excel_path, output_sheet_name, student, error_data, subjects)

    finally:
        driver.quit()
        print(f"\n所有考生处理完毕，结果已保存至：{os.path.abspath(output_excel_path)}")


if __name__ == "__main__":
    main()